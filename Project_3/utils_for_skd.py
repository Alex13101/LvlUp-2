import openpyxl


def get_search(message, quantity):
    excel_file = openpyxl.load_workbook('skl_b1.xlsx')
    sheet_sheet = excel_file['Лист1']  # Выбираем лист в файле exel
    found = []
    quantity = []

    search = message.text

    for x in range(1, sheet_sheet.max_row + 1):  # Печать всех значений из столбца

        if search in sheet_sheet.cell(row=x, column=2).value:
            found = sheet_sheet.cell(row=x, column=2).value
            quantity = sheet_sheet.cell(row=x, column=3).value
            # print(f" {found} имеется в количестве {quantity} шт. ")
    excel_file.close()

    return found, quantity
    # print("Not found!")


def is_user(message):
    excel_file = openpyxl.load_workbook('users_skd.xlsx')
    user_sheet = excel_file['Лист1']  # Выбираем лист в файле exel
    search = message.text
    z = 0

    for x in range(1, user_sheet.max_row + 1):  # Печать всех значений из столбца
        found = user_sheet.cell(row=x, column=1).value
        if search == user_sheet.cell(row=x, column=1).value:
            us = user_sheet.cell(row=x, column=2).value
            z += 1

    if z > 0:
        return True
    else:
        return False





  # us = {'1234': 'Алексей', '4321': 'Антон', '231': 'Жорик'}
    # for i, v in us.items():  # Идем в цикле по ключам и значениям
    #     if i == us.items():
    #         bot.send_message(message.chat.id,
    #                          f" Привет! {v}Что будем искать?")

            #@bot.message_handler(content_types=["text"])
