import openpyxl
import telebot
from tqdm import tqdm
from utils_for_skd import is_user
from telebot import types
from datetime import date

current_date = date.today()

bot = telebot.TeleBot('_____________________')  # Токен от bot_father
excel_file = openpyxl.load_workbook('skl_b1.xlsx')  # Открытие exel-файла


@bot.message_handler(commands=['help'])
def register(message):
    bot.send_message(message.chat.id, 'Поиск информации по складам. Нажмите: /start')


@bot.message_handler(commands=['start'])
def handle_text(message):
    bot.send_message(message.chat.id, 'Введите пароль! ')
    bot.register_next_step_handler(message, password_chek)


def password_chek(message):
    user = is_user(message)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Склад_В1")
    item2 = types.KeyboardButton("Склад_В2")
    item3 = types.KeyboardButton("Склад_В3")
    markup.add(item1)
    markup.add(item2)
    markup.add(item3)
    if user is not None:
        bot.send_message(message.chat.id, f'Привет {user} Где будем искать? Нажмите: ', reply_markup=markup)

    else:
        bot.send_message(message.chat.id, " Попробуйте еще раз: /start")


@bot.message_handler(content_types=["text"])
def redirection(message):
    if message.text == "Склад_В1":
        markup1 = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Изменить позицию на В1")
        item2 = types.KeyboardButton("Выйти в главное меню")
        item3 = types.KeyboardButton("Склад_В1")
        markup1.add(item1)
        markup1.add(item2)
        markup1.add(item3)
        bot.send_message(message.chat.id, 'Что будем искать? ', reply_markup=markup1)
        sheet_sheet = excel_file['Лист1']
        bot.register_next_step_handler(message, search_exel_b1, excel_file, sheet_sheet)
    elif message.text == "Склад_В2":
        markup1 = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Изменить позицию на В2")
        item2 = types.KeyboardButton("Выйти в главное меню")
        item3 = types.KeyboardButton("Склад_В2")
        markup1.add(item1)
        markup1.add(item2)
        markup1.add(item3)
        bot.send_message(message.chat.id, 'Что будем искать? ', reply_markup=markup1)
        sheet_sheet = excel_file['Лист2']
        bot.register_next_step_handler(message, search_exel_b1, sheet_sheet)
    elif message.text == "Склад_В3":
        markup1 = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Изменить позицию на В3")
        item2 = types.KeyboardButton("Выйти в главное меню")
        item3 = types.KeyboardButton("Склад_В3")
        # item6 = types.KeyboardButton("Продолжить на этом складе")
        markup1.add(item1)
        markup1.add(item2)
        markup1.add(item3)
        sheet_sheet = excel_file['Лист3']
        bot.send_message(message.chat.id, 'Что будем искать? ', reply_markup=markup1)
        bot.register_next_step_handler(message, search_exel_b1, sheet_sheet)
    elif message.text == "Выйти в главное меню":
        bot.send_message(message.chat.id, f'Введите пароль:')
        bot.register_next_step_handler(message, password_chek)
    elif message.text == "Изменить позицию на В1":
        bot.send_message(message.chat.id, f'Введите номер изменяемой позиции: ')
        sheet_sheet = excel_file['Лист1']
        bot.register_next_step_handler(message, save_number_pos_b1, sheet_sheet)
    elif message.text == "Изменить позицию на В2":
        bot.send_message(message.chat.id, f'Введите номер изменяемой позиции: ')
        sheet_sheet = excel_file['Лист2']
        bot.register_next_step_handler(message, save_number_pos_b1, sheet_sheet)
    elif message.text == "Изменить позицию на В3":
        bot.send_message(message.chat.id, f'Введите номер изменяемой позиции: ')
        sheet_sheet = excel_file['Лист3']
        bot.register_next_step_handler(message, save_number_pos_b1, sheet_sheet)
    else:
        bot.send_message(message.chat.id, 'Выберите команду из списка')


def search_exel_b1(message, sheet_sheet):  # Поиск для складов общий
    search = message.text
    for x in tqdm(range(1, sheet_sheet.max_row + 1)):  # Печать всех значений из столбца
        #found = sheet_sheet.cell(row=x, column=2).value
        if search in sheet_sheet.cell(row=x, column=2).value:
            #quantity = sheet_sheet.cell(row=x, column=3).value
            #print(f" {found} имеется в количестве {quantity} шт. ")
            bot.send_message(message.chat.id,
                             f" поз.№ {sheet_sheet.cell(row=x, column=1).value} - {sheet_sheet.cell(row=x, column=2).value} в наличии {sheet_sheet.cell(row=x, column=3).value} шт.")

        elif x == sheet_sheet.max_row:
            print("Not found!")
            bot.send_message(message.chat.id, " Search complited. \n Что-то еще? ")


def save_number_pos_b1(message, sheet_sheet):  # Запись значения в файл для склада_В1 в формате (int)
    try:
        pos = int(message.text)
        bot.send_message(message.chat.id, " Какое количество записать? ")
        bot.register_next_step_handler(message, save_quantity_b1, sheet_sheet, pos)
    except:  # Проверка на ошибку по целому числу int
        bot.send_message(message.chat.id, " Необходимо числом. Выбирай что делать! ")


def save_quantity_b1(message, sheet_sheet, pos):
    # excel_file = openpyxl.load_workbook('skl_b1.xlsx')  # Открытие exel-файла
    if pos <= sheet_sheet.max_row+1:
        print(sheet_sheet.max_row+1)
        #excel_file = openpyxl.load_workbook('skl_b1.xlsx')
        quant = message.text
        sheet_sheet.cell(row=pos + 1, column=3).value = quant
        sheet_sheet.cell(row=pos + 1, column=4).value = current_date
        bot.send_message(message.chat.id, 'Готово!')
        bot.send_message(message.chat.id, 'Выберите команду из списка')
        excel_file.save('skl_b1.xlsx')

    else:
        bot.send_message(message.chat.id, 'Такой позиции нет! Давайте еще раз')


bot.polling(none_stop=True, interval=0)

