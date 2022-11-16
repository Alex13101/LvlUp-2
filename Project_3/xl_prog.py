import openpyxl

excel_file = openpyxl.load_workbook('skl_b1.xlsx')
#
# # sheet names
# print(excel_file.sheetnames)

sheet_sheet = excel_file['Лист1'] #Выбираем лист в файле exel
#sheet_sheet = excel_file.active # Выбираем активный лист в exel-файле
cell_obj = sheet_sheet.cell(row=3, column=2) # Координата в активном листе
print(type(cell_obj)) # Неизвестный
print(f'Employees[A1]={cell_obj.value}')

# second way
print(f'Employees[A1]={sheet_sheet["B1"].value}') # Так же координата
print(f'Total Rows = {sheet_sheet.max_row} and Total Columns = {sheet_sheet.max_column}')

#header_cells_generator = sheet_sheet.iter_rows(max_row=1)  #Функция ITER_ROWS () генерирует ячейки из рабочего листа по строке.
                                                            # Мы можем использовать его,
                                                            # чтобы получить клетки из определенного ряда.
# for header_cells_tuple in header_cells_generator:
#     for i in range(len(header_cells_tuple)):
#         print(header_cells_tuple[i].value)
#
# for x in range(1, sheet_sheet.max_row+1):    #Печать всех значений из столбца
#     print(sheet_sheet.cell(row=x, column=2).value)
#
# for x in range(2, sheet_sheet.max_column+1):   #Печать всех значений из строки
#     print(sheet_sheet.cell(row=3, column=x).value)

# cells = sheet_sheet['A2':'C3']
# for id, name, role in cells:     #Мы можем пройти ассортимент ячеек для чтения
#                                     # нескольких ячеек одновременно.
#     print(f'Employee[{id.value}, {name.value}, {role.value}]')

for row in sheet_sheet.iter_rows(min_row=2, min_col=2, max_row=4, max_col=3):#Итализация клетки по рядам
    for cell in row:
        print(cell.value, end="|")
    print("")

for col in sheet_sheet.iter_cols(min_row=2, min_col=2, max_row=4, max_col=3):#Итализаторы клетки по столбцам
    for cell in col:
        print(cell.value, end="|")
    print("")

